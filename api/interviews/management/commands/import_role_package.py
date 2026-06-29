from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from api.core.constants import ExpectedAnswerType, InterviewEvaluationTier, QuestionLifecycleStatus
from api.interviews.models import InterviewConfiguration, InterviewRubric
from api.questions.models import QuestionTemplate

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - environment/setup fallback
    raise ImportError(
        "openpyxl is required to import MeritLense role packages. "
        "Install project dependencies first."
    ) from exc


LANGUAGE_MAP = {
    "en": "EN",
    "ar": "AR",
}

TIER_MAP = {
    "full": InterviewEvaluationTier.FULL,
    "screening": InterviewEvaluationTier.SCREENING,
    "both": InterviewEvaluationTier.BOTH,
}


class Command(BaseCommand):
    help = "Import a MeritLense question bank role package workbook into interview configs, rubrics, and question templates."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Absolute or relative path to the role package workbook (.xlsx)")

    @transaction.atomic
    def handle(self, *args, **options):
        workbook_path = Path(options["xlsx_path"]).expanduser().resolve()
        if not workbook_path.exists():
            raise CommandError(f"Workbook not found: {workbook_path}")

        workbook = load_workbook(workbook_path, data_only=True)
        metadata = self._parse_metadata(workbook)
        rubric_rows = self._parse_rubrics(workbook)
        question_rows = self._parse_questions(workbook)
        criteria_by_skill = self._parse_answer_blueprint(workbook)
        if question_rows:
            metadata["rubric_version"] = question_rows[0]["rubric_version"]
            metadata["question_set_version"] = question_rows[0]["question_set_version"]
        else:
            metadata["rubric_version"] = ""
            metadata["question_set_version"] = ""

        self._sync_rubrics(metadata, rubric_rows, criteria_by_skill)
        self._sync_questions(metadata, question_rows)
        self._sync_configs(metadata, question_rows)

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported role package for {metadata['role_name']} ({metadata['role_code']}) from {workbook_path.name}"
            )
        )

    def _parse_metadata(self, workbook):
        sheet = workbook["README"]
        return {
            "role_name": sheet["B3"].value,
            "role_code": sheet["B4"].value,
            "package_version": sheet["B5"].value if sheet["B5"].value else sheet["B7"].value,
            "languages": sheet["B6"].value if sheet["B6"].value else sheet["B8"].value,
            "rubric_version": sheet["B7"].value or "",
            "question_set_version": sheet["B8"].value or "",
        }

    def _parse_rubrics(self, workbook):
        sheet = workbook["Rubric_Map"]
        rows = []
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row[0] or row[0] == "TOTAL":
                continue
            weight_value = row[2]
            try:
                if isinstance(weight_value, str) and weight_value.endswith("%"):
                    weight = Decimal(weight_value.rstrip("%")) / Decimal("100")
                else:
                    weight = Decimal(str(weight_value))
                max_score = int(row[3])
            except (InvalidOperation, TypeError, ValueError):
                continue
            rows.append(
                {
                    "skill_tag": row[0],
                    "scoring_category": row[1],
                    "weight": weight,
                    "max_score": max_score,
                    "scoring_type": row[4],
                    "domain": row[5] or "",
                    "notes": row[6] or "",
                }
            )
        return rows

    def _parse_questions(self, workbook):
        rows = []
        for sheet_name in ["Questions_EN", "Questions_AR"]:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=3, values_only=True):
                if self._is_v12_question_row(row):
                    rows.append(
                        {
                            "question_code": row[0] or "",
                            "question_version": row[1] or "",
                            "question_status": (row[2] or QuestionLifecycleStatus.ACTIVE).lower(),
                            "sequence_number": row[3],
                            "domain": row[4] or "",
                            "skill_tag": row[5] or "",
                            "question_text": row[6] or "",
                            "question_type": row[7] or "",
                            "question_format": row[8] or "TEXT",
                            "difficulty": row[9] or "MEDIUM",
                            "difficulty_score": int(row[10] or 2),
                            "scoring_type": row[11] or "",
                            "weight": Decimal(str(row[12])),
                            "evaluation_tier": TIER_MAP[str(row[13]).lower()],
                            "estimated_time_seconds": int(row[14] or 30),
                            "is_mandatory": self._as_bool(row[15], default=True),
                            "expected_answer_type": row[16] or ExpectedAnswerType.STRUCTURED,
                            "follow_up_allowed": self._as_bool(row[17], default=False),
                            "critical_question": self._as_bool(row[18], default=False),
                            "language": LANGUAGE_MAP.get(str(row[19]).lower(), str(row[19]).upper()),
                            "rubric_version": workbook["README"]["B7"].value or "",
                            "question_set_version": workbook["README"]["B8"].value or "",
                        }
                    )
                    continue
                if not isinstance(row[0], int):
                    continue
                rows.append(
                    {
                        "question_code": "",
                        "question_version": "1.0",
                        "question_status": QuestionLifecycleStatus.ACTIVE,
                        "sequence_number": row[0],
                        "domain": row[1] or "",
                        "skill_tag": row[2] or "",
                        "question_text": row[3] or "",
                        "question_type": "scenario" if row[4] == "SCENARIO" else "knowledge",
                        "question_format": row[4] or "TEXT",
                        "difficulty": row[5] or "MEDIUM",
                        "difficulty_score": self._difficulty_score(row[5] or "MEDIUM"),
                        "scoring_type": row[6] or "",
                        "weight": Decimal(str(row[7])),
                        "evaluation_tier": TIER_MAP[str(row[8]).lower()],
                        "estimated_time_seconds": 30,
                        "is_mandatory": True,
                        "expected_answer_type": ExpectedAnswerType.STRUCTURED,
                        "follow_up_allowed": False,
                        "critical_question": False,
                        "language": LANGUAGE_MAP.get(str(row[9]).lower(), str(row[9]).upper()),
                        "rubric_version": row[10] or "",
                        "question_set_version": row[11] or "",
                    }
                )
        return rows

    def _parse_answer_blueprint(self, workbook):
        sheet = workbook["Answer_Blueprint"]
        grouped = defaultdict(list)
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row[0]:
                continue
            if len(row) >= 8:
                grouped[row[1]].append(
                    {
                        "question_ref": row[0],
                        "question_type": row[2],
                        "expected_answer_type": row[3],
                        "must_include_points": row[4],
                        "ideal_keywords": row[5],
                        "negative_indicators": row[6],
                        "score_note": row[7],
                    }
                )
            else:
                grouped[row[1]].append(
                    {
                        "question_ref": row[2],
                        "must_include_points": row[3],
                        "ideal_keywords": row[4],
                        "negative_indicators": row[5],
                        "score_note": row[6],
                    }
                )
        return grouped

    def _sync_rubrics(self, metadata, rubric_rows, criteria_by_skill):
        for row in rubric_rows:
            InterviewRubric.objects.update_or_create(
                role_code=metadata["role_code"],
                skill_tag=row["skill_tag"],
                rubric_version=metadata["rubric_version"],
                defaults={
                    "role_name": metadata["role_name"],
                    "role_code": metadata["role_code"],
                    "scoring_category": row["scoring_category"],
                    "weight": row["weight"],
                    "max_score": row["max_score"],
                    "scoring_type": row["scoring_type"],
                    "domain": row["domain"],
                    "notes": row["notes"],
                    "question_set_version": metadata["question_set_version"],
                    "evaluation_criteria": criteria_by_skill.get(row["skill_tag"], []),
                    "is_active": True,
                },
            )

    def _sync_questions(self, metadata, question_rows):
        QuestionTemplate.objects.filter(role_code=metadata["role_code"]).exclude(
            question_set_version=metadata["question_set_version"]
        ).update(is_active=False, question_status=QuestionLifecycleStatus.ARCHIVED)

        for row in question_rows:
            QuestionTemplate.objects.update_or_create(
                role_code=metadata["role_code"],
                question_code=row["question_code"],
                language=row["language"],
                question_version=row["question_version"],
                defaults={
                    "role_name": metadata["role_name"],
                    "domain": row["domain"],
                    "skill_tag": row["skill_tag"],
                    "skill": row["skill_tag"],
                    "sequence_number": row["sequence_number"],
                    "question_text": row["question_text"],
                    "question_type": row["question_type"],
                    "question_format": row["question_format"],
                    "question_status": row["question_status"],
                    "difficulty": row["difficulty"],
                    "difficulty_score": row["difficulty_score"],
                    "scoring_type": row["scoring_type"],
                    "weight": row["weight"],
                    "estimated_time_seconds": row["estimated_time_seconds"],
                    "expected_answer_type": row["expected_answer_type"],
                    "evaluation_tier": row["evaluation_tier"],
                    "rubric_version": row["rubric_version"],
                    "question_set_version": row["question_set_version"],
                    "is_mandatory": row["is_mandatory"],
                    "follow_up_allowed": row["follow_up_allowed"],
                    "critical_question": row["critical_question"],
                    "is_active": row["question_status"] == QuestionLifecycleStatus.ACTIVE,
                },
            )

    def _sync_configs(self, metadata, question_rows):
        grouped = defaultdict(int)
        rubric_versions = {}
        question_set_versions = {}
        for row in question_rows:
            effective_tiers = self._effective_config_tiers(row["evaluation_tier"])
            for effective_tier in effective_tiers:
                key = (row["language"], effective_tier)
                grouped[key] += 1
                rubric_versions[key] = row["rubric_version"]
                question_set_versions[key] = row["question_set_version"]

        for (language, evaluation_tier), total_questions in grouped.items():
            InterviewConfiguration.objects.update_or_create(
                role_code=metadata["role_code"],
                language=language,
                evaluation_tier=evaluation_tier,
                defaults={
                    "role_name": metadata["role_name"],
                    "duration_minutes": 30 if evaluation_tier == InterviewEvaluationTier.SCREENING else 45,
                    "total_questions": total_questions,
                    "allow_retries": True,
                    "max_retries": 1,
                    "enable_translation": language == "AR",
                    "enable_task_module": evaluation_tier == InterviewEvaluationTier.FULL,
                    "enable_integrity_checks": evaluation_tier == InterviewEvaluationTier.FULL,
                    "rubric_version": rubric_versions[(language, evaluation_tier)],
                    "question_set_version": question_set_versions[(language, evaluation_tier)],
                    "is_active": True,
                },
            )

    def _effective_config_tiers(self, evaluation_tier):
        if evaluation_tier == InterviewEvaluationTier.BOTH:
            return [InterviewEvaluationTier.FULL, InterviewEvaluationTier.SCREENING]
        return [evaluation_tier]

    def _is_v12_question_row(self, row):
        return len(row) >= 20 and row[0] and isinstance(row[3], int)

    def _as_bool(self, value, default=False):
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip().upper() == "TRUE"

    def _difficulty_score(self, difficulty):
        return {
            "EASY": 1,
            "MEDIUM": 2,
            "HARD": 3,
        }.get(str(difficulty).upper(), 2)
